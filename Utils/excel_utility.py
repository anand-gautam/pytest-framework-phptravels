import openpyxl

def get_row_count(xl_path, sheet_name):
    workbook = openpyxl.load_workbook(xl_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_col_count(xl_path, sheet_name):
    workbook = openpyxl.load_workbook(xl_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_cell_data(xl_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(xl_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=col_num).value


def write_data(xl_path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(xl_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(xl_path)
